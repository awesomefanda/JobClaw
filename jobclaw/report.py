"""Phase 5 — REPORT: Generate the final Excel output.

Reads data/enriched.json and produces a single .xlsx with 4 sheets:
  1. Job Matches (ranked by action_score, all data in one row)
  2. Hiring Posts (LinkedIn "I'm hiring" feed)
  3. Salary Data (Levels.fyi + Blind per company)
  4. Pipeline (tracker for user to update as they progress)
"""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from jobclaw.logger import get_logger

log = get_logger("report")

# Known company name aliases → canonical name
_CO_ALIASES: dict[str, str] = {
    "amazon web services": "Amazon",
    "aws": "Amazon",
    "alphabet inc": "Google",
    "alphabet": "Google",
    "meta platforms": "Meta",
    "facebook": "Meta",
    "x corp": "X (Twitter)",
    "twitter": "X (Twitter)",
    "microsoft corporation": "Microsoft",
    "apple inc": "Apple",
    "salesforce inc": "Salesforce",
    "servicenow inc": "ServiceNow",
}

_CO_SUFFIXES = (" inc.", " inc", " llc", " ltd", " corp.", " corp", " corporation", " co.", " co")


def _normalize_co(name: str) -> str:
    """Normalize company name for deduplication (Amazon / AWS → Amazon)."""
    stripped = name.strip()
    lower = stripped.lower()
    if lower in _CO_ALIASES:
        return _CO_ALIASES[lower]
    for alias, canonical in _CO_ALIASES.items():
        if lower == alias:
            return canonical
    for sfx in _CO_SUFFIXES:
        if lower.endswith(sfx):
            stripped = stripped[:len(stripped) - len(sfx)].strip()
            break
    return stripped

DATA = Path(__file__).resolve().parent.parent / "data"
ENRICHED_JSON = DATA / "enriched.json"
REPORTS = DATA / "reports"

# ── Styles ─────────────────────────────────────────────────────
H_FILL  = PatternFill('solid', fgColor='1B2A4A')
H_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
A_FILL  = PatternFill('solid', fgColor='E8F5E9')  # green
B_FILL  = PatternFill('solid', fgColor='FFF3E0')  # orange
BODY    = Font(name='Arial', size=9)
BOLD    = Font(name='Arial', bold=True, size=9)
HI_FONT = Font(name='Arial', bold=True, color='1B7A2B', size=10)
MID_FONT = Font(name='Arial', color='B8860B', size=10)
WRAP    = Alignment(wrap_text=True, vertical='top')
BDR     = Border(bottom=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='E8E8E8'))


def _header_row(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill = H_FONT, H_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 34


def _cell(ws, row, col, val, font=BODY, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font, c.alignment, c.border = font, WRAP, BDR
    if fill:
        c.fill = fill
    return c


def generate_report(enriched: list[dict] | None = None) -> str:
    """Generate Excel report. Returns filepath."""
    log.info("=" * 50)
    log.info("PHASE 5: REPORT")
    log.info("=" * 50)

    REPORTS.mkdir(exist_ok=True)

    if enriched is None:
        if not ENRICHED_JSON.exists():
            log.error("No enriched.json — cannot generate report")
            return ""
        enriched = json.loads(ENRICHED_JSON.read_text())

    if not enriched:
        log.warning("No matches to report")
        return ""

    wb = Workbook()

    # ── Sheet 1: Job Matches ───────────────────────────────────
    ws = wb.active
    ws.title = "Job Matches"
    headers = [
        "Rank", "Action\nScore", "Fit\nScore", "Track", "Company", "Title",
        "Location", "Salary\n(Levels.fyi)", "Blind\nOffers",
        "Company\nHealth", "Your\nConnections", "Hiring\nPosts",
        "Best Contact", "Outreach Draft", "Apply URL", "Source", "Notes"
    ]
    widths = [5, 7, 6, 6, 14, 28, 12, 16, 24, 24, 24, 28, 20, 55, 28, 10, 22]
    _header_row(ws, headers, widths)

    for idx, ej in enumerate(enriched):
        row = idx + 2
        track = ej.get("track", "B")
        fill = A_FILL if track == "A" else B_FILL
        signals = ej.get("signals", {})

        # Connections
        conns = ej.get("my_connections", [])
        conn_text = "\n".join(f"{c['name']} ({c['position']})" for c in conns[:3]) or "—"

        # Hiring posts (LinkedIn Playwright scraper)
        hp = signals.get("hiring_posts", [])
        hp_parts = [f"{p.get('poster','?')}: {p.get('snippet','')[:120]}\n{p.get('url','')}" for p in hp[:3]]
        hp_text = "\n\n".join(hp_parts) or "—"

        # Blind offers
        blind = signals.get("blind_offers", [])
        blind_text = "\n".join(o[:150] for o in blind[:3]) or "—"

        # Company health
        layoffs = signals.get("layoffs", {})
        sentiment = signals.get("blind_sentiment", {})
        funding = signals.get("funding", "")
        salary = signals.get("salary", {})
        health_parts = []
        if layoffs.get("had_layoffs"):
            health_parts.append(f"⚠️ Layoffs: {layoffs.get('detail','')[:80]}")
        else:
            health_parts.append("✅ No recent layoffs")
        if funding:
            health_parts.append(f"💰 {funding[:80]}")
        if sentiment.get("red_flags"):
            for n in sentiment.get("negative", [])[:1]:
                health_parts.append(f"⚠️ {n[:80]}")
        health_text = "\n".join(health_parts)

        # Salary — levels.fyi range, then job listing range, then offer submission hint
        sal_text = salary.get("tc_range", "") or salary.get("base_range", "")
        if not sal_text:
            sm, sx = ej.get("salary_min", ""), ej.get("salary_max", "")
            if sm and sx:
                sal_text = f"${sm} - ${sx}"
        if not sal_text:
            raw_offers = signals.get("levels_offers", [])
            for o in raw_offers:
                tc = o.get("tc_range", "") if isinstance(o, dict) else ""
                if tc:
                    sal_text = tc
                    break
        if not sal_text:
            n = len(signals.get("levels_offers", []))
            sal_text = f"📊 {n} recent submissions" if n else "—"

        # Best contact
        bc = ej.get("best_contact", {})
        bc_parts = []
        if bc.get("name"):
            bc_parts.append(f"{bc['name']} ({bc.get('title','')})")
        if bc.get("linkedin_url"):
            bc_parts.append(bc["linkedin_url"])
        if bc.get("email"):
            bc_parts.append(f"📧 {bc['email']}")
        src = bc.get("source", "")
        if src == "your_connection":
            bc_parts.append("⭐ YOUR CONNECTION")
        elif src == "founder_direct":
            bc_parts.append("🎯 DIRECT EMAIL")
        bc_text = "\n".join(bc_parts) or "—"

        # Notes
        notes = ""
        if track == "A":
            notes = f"Track A: {bc.get('source','').replace('_',' ')}"
        else:
            notes = "Track B: no warm path — apply with tailored resume"

        values = [
            idx + 1, ej.get("action_score", 0), ej.get("fit_score", 0),
            f"Track {track}", ej.get("company", ""), ej.get("title", ""),
            ej.get("location", ""), sal_text, blind_text,
            health_text, conn_text, hp_text, bc_text,
            ej.get("outreach_draft", ""), ej.get("job_url", ""),
            ej.get("source", ""), notes,
        ]

        for col, val in enumerate(values, 1):
            font = BODY
            if col == 2:
                font = HI_FONT if ej.get("action_score", 0) >= 1.0 else MID_FONT
            elif col == 3:
                font = HI_FONT if ej.get("fit_score", 0) >= 0.85 else MID_FONT
            elif col in (4, 5):
                font = BOLD
            _cell(ws, row, col, val, font=font, fill=fill)

        ws.row_dimensions[row].height = 140

    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f"A1:Q{len(enriched)+1}"

    # ── Sheet 2: Hiring Posts ──────────────────────────────────
    ws2 = wb.create_sheet("Hiring Posts")
    _header_row(ws2, ["Person / Poster", "Type", "Company", "Snippet", "URL"], [22, 12, 14, 55, 35])
    row = 2
    seen_urls = set()
    for ej in enriched:
        sig2 = ej.get("signals", {})
        # Text posts (site:linkedin.com/posts)
        for p in sig2.get("hiring_posts", []):
            if p.get("url") in seen_urls:
                continue
            seen_urls.add(p.get("url"))
            for col, val in enumerate([p.get("poster",""), "Post", ej.get("company",""), p.get("snippet",""), p.get("url","")], 1):
                _cell(ws2, row, col, val)
            ws2.row_dimensions[row].height = 50
            row += 1
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: Salary Data ───────────────────────────────────
    ws3 = wb.create_sheet("Salary Data")
    _header_row(ws3,
        ["Company", "TC Range\n(Levels.fyi)", "Base Range\n(Levels.fyi)", "Recent Submissions\n(= actively interviewing)", "Blind Offers", "Funding"],
        [14, 18, 18, 50, 45, 40])
    seen_co: set[str] = set()
    row = 2
    for ej in enriched:
        co_raw = ej.get("company", "")
        co = _normalize_co(co_raw)
        if co.lower() in seen_co:
            continue
        seen_co.add(co.lower())
        sig = ej.get("signals", {})
        sal = sig.get("salary", {})

        # levels_offers — list of {snippet, tc_range} dicts (new format) or legacy strings
        raw_offers = sig.get("levels_offers", [])
        offer_lines = []
        for o in raw_offers[:3]:
            if isinstance(o, dict):
                line = o.get("snippet", "")[:150]
                tc = o.get("tc_range", "")
                if tc:
                    line = f"💰 {tc} — {line}"
            else:
                line = str(o)[:150]
            if line:
                offer_lines.append(line)
        submissions_text = "\n\n".join(offer_lines) if offer_lines else "—"

        blind = "\n".join(sig.get("blind_offers", [])[:2]) or "—"
        for col, val in enumerate([
            co,
            sal.get("tc_range", "") or "—",
            sal.get("base_range", "") or "—",
            submissions_text,
            blind,
            sig.get("funding", "") or "—",
        ], 1):
            _cell(ws3, row, col, val)
        ws3.row_dimensions[row].height = 70
        row += 1
    ws3.freeze_panes = 'A2'

    # ── Sheet 4: Pipeline ──────────────────────────────────────
    ws4 = wb.create_sheet("Pipeline")
    _header_row(ws4, ["#","Company","Title","Track","Action\nScore","Status","Outreach\nSent","Response","Applied","Interview","Outcome","Notes"],
                [5,14,28,6,7,12,12,12,12,12,12,25])
    for idx, ej in enumerate(enriched):
        row = idx + 2
        for col, val in enumerate([idx+1, ej.get("company",""), ej.get("title",""), ej.get("track",""), ej.get("action_score",0), "New", "","","","","",""], 1):
            _cell(ws4, row, col, val)
    ws4.freeze_panes = 'C2'

    # ── Sheet 5: Alerts ────────────────────────────────────────
    ws5 = wb.create_sheet("🔔 Alerts")
    _header_row(ws5,
        ["#", "Company", "Title", "Alert Signals", "Action\nScore", "Fit\nScore", "Track", "Best Contact", "Apply URL"],
        [5, 14, 28, 36, 8, 8, 6, 28, 35])

    ALERT_FILL = PatternFill('solid', fgColor='FFF9C4')  # light yellow
    HOT_FILL   = PatternFill('solid', fgColor='FFCDD2')  # light red

    alerts = []
    for job in enriched:
        signals = job.get("signals", {})
        reasons = []

        if job.get("action_score", 0) >= 1.2:
            reasons.append("🔥 Top Pick")
        if job.get("fit_score", 0) >= 0.85:
            reasons.append("🎯 Strong Fit")
        if signals.get("hiring_posts"):
            reasons.append("📢 Active Hiring")
        if signals.get("funding"):
            reasons.append("💰 Recently Funded")
        if job.get("track") == "A":
            reasons.append("⭐ Warm Path")

        date_str = job.get("date_posted", "")
        if date_str:
            try:
                posted = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
                if (datetime.now() - posted).days <= 7:
                    reasons.append("📅 Fresh (<7 days)")
            except Exception:
                pass

        if job.get("action_score", 0) >= 1.0 or len(reasons) >= 2:
            alerts.append((job, reasons))

    alerts.sort(key=lambda x: x[0].get("action_score", 0), reverse=True)

    for idx, (job, reasons) in enumerate(alerts):
        row = idx + 2
        fill = HOT_FILL if job.get("action_score", 0) >= 1.2 else ALERT_FILL

        bc = job.get("best_contact", {})
        bc_parts = []
        if bc.get("name"):
            bc_parts.append(f"{bc['name']} ({bc.get('title', '')})")
        if bc.get("source") == "your_connection":
            bc_parts.append("⭐ YOUR CONNECTION")
        elif bc.get("source") == "founder_direct":
            bc_parts.append("🎯 DIRECT EMAIL")
        bc_text = "\n".join(bc_parts) or "—"

        values = [
            idx + 1,
            job.get("company", ""),
            job.get("title", ""),
            "\n".join(reasons),
            job.get("action_score", 0),
            job.get("fit_score", 0),
            f"Track {job.get('track', 'B')}",
            bc_text,
            job.get("job_url", ""),
        ]
        for col, val in enumerate(values, 1):
            font = BOLD if col in (2, 3) else BODY
            if col == 5:
                font = HI_FONT if job.get("action_score", 0) >= 1.0 else MID_FONT
            elif col == 6:
                font = HI_FONT if job.get("fit_score", 0) >= 0.85 else MID_FONT
            _cell(ws5, row, col, val, font=font, fill=fill)
        ws5.row_dimensions[row].height = 70

    ws5.freeze_panes = 'C2'
    log.info(f"   Alerts: {len(alerts)} jobs flagged")

    # ── Save ───────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = str(REPORTS / f"jobclaw_report_{ts}.xlsx")
    wb.save(filepath)

    log.info(f"📊 Report saved: {filepath}")
    log.info(f"   {len(enriched)} matches | Track A: {sum(1 for j in enriched if j.get('track')=='A')} | Track B: {sum(1 for j in enriched if j.get('track')=='B')}")
    return filepath
